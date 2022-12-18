from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule
import datetime


def export(currency_data):
    wb = Workbook()
    ws = wb.active
    ws.title = datetime.date.today().strftime("%m-%d-%Y")

    create_header(ws)

    for d in currency_data:
        append_data_row(d, ws)

    scale_rule = ColorScaleRule(start_type='percentile', start_value=-0, start_color='00FF0000',
                                mid_type='percentile', mid_value=50, mid_color='00FFFF00',
                                end_type='percentile', end_value=100, end_color='0000FF00')
    ws.conditional_formatting.add(f"E2:E{len(currency_data) + 1}", scale_rule)

    wb.save("Currencies/Currencies.xlsx")
    print('Excel export completed')


def append_data_row(d, ws):
    symbol = d.symbol.replace("=X", "")
    name = d.name
    last_price = float(d.last_price.replace(",", ""))
    change = float(d.change.replace(",", ""))
    percent_change = float(d.percent_change.replace("%", "").replace(",", ""))

    ws.append([symbol, name, last_price, change, percent_change])


def create_header(ws):
    header_font = Font(bold=True)
    ws['A1'].font = header_font
    ws['A1'] = 'Symbol'
    ws['B1'].font = header_font
    ws['B1'] = 'Name'
    ws['C1'].font = header_font
    ws['C1'] = 'Last Price'
    ws['D1'].font = header_font
    ws['D1'] = 'Change'
    ws['E1'].font = header_font
    ws['E1'] = '% Change'

    # AFAICT the column width conversion depends on default system font and maybe screen resolution.
    # But setting width approach would be something like this.
    ws.column_dimensions['A'].width = 10.00
    ws.column_dimensions['B'].width = 10.00
    ws.column_dimensions['C'].width = 10.00
    ws.column_dimensions['D'].width = 10.00
    ws.column_dimensions['E'].width = 10.00
